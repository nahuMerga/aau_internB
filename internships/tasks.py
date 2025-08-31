from celery import shared_task
from django.db import transaction
from openpyxl import load_workbook
from internships.models import ThirdYearStudentList
from advisors.models import Advisor
from advisors.tasks import notify_advisor_task
from utils.generate_email import generate_email
from utils.get_next_available_advisor import get_next_available_advisor
import os

@shared_task(bind=True, max_retries=3)
def process_student_excel_task(self, file_path):
    """Process student Excel file asynchronously using openpyxl"""
    try:
        # Load workbook using openpyxl
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        
        # Get headers
        headers = [cell.value for cell in sheet[1]]
        required_columns = {'university_id', 'full_name'}
        
        if not all(col in headers for col in required_columns):
            raise ValueError(f"Missing columns. Required: {required_columns}")

        created_students = []
        
        with transaction.atomic():
            # Iterate through rows (starting from row 2 to skip headers)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                    
                # Create a dictionary from row values and headers
                row_data = dict(zip(headers, row))
                
                university_id = str(row_data['university_id']).strip()[:20]
                full_name = str(row_data['full_name']).strip()

                if ThirdYearStudentList.objects.filter(university_id=university_id).exists():
                    continue

                email = generate_email(full_name, university_id)
                advisor = get_next_available_advisor()
                
                if not advisor:
                    raise ValueError("No available advisor.")

                student = ThirdYearStudentList.objects.create(
                    university_id=university_id,
                    full_name=full_name,
                    institutional_email=email,
                    assigned_advisor=advisor
                )

                created_students.append({
                    "university_id": university_id,
                    "full_name": full_name,
                    "email": email,
                    "advisor": advisor.first_name
                })

        # Notify all advisors about their new students
        for advisor in Advisor.objects.all():
            notify_advisor_task.delay(advisor.id)

        # Clean up temporary file
        if os.path.exists(file_path):
            os.unlink(file_path)

        return {
            "message": f"{len(created_students)} students uploaded successfully.",
            "data": created_students
        }
    except Exception as e:
        # Clean up temporary file on error
        if os.path.exists(file_path):
            os.unlink(file_path)
        self.retry(countdown=60, exc=e)
